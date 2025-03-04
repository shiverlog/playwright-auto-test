# from asyncio.windows_events import NULL
# from multiprocessing.connection import wait
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import json


class ToExcel():
    def __init__(self):
        self.export_excel_data()

    def export_excel_data(self):
        with open('common/속도측정.json') as f:
            self.data=json.load(f)
            f.close()

        wb = Workbook()
        ws = wb.active
        ws['C1'] = '1회'
        ws['D1'] = '2회'
        ws['E1'] = '3회'
        ws['F1'] = '4회'
        ws['G1'] = '5회'
        # ws['H1'] = '평균'
        count = 1
        agencys = ['lg','kt','skt']
        for agency in agencys:
            ws[f'B{count}'] = agency
            count+=1
            for key,value in  self.data.get(agency).items():
                ws[f'B{count}'] = key
                측정리스트 = value.get("측정")
                ws[f'C{count}'] = 측정리스트[0].get("time") 
                ws[f'D{count}'] = 측정리스트[1].get("time")
                ws[f'E{count}'] = 측정리스트[2].get("time")
                ws[f'F{count}'] = 측정리스트[3].get("time")
                ws[f'G{count}'] = 측정리스트[4].get("time")
                ws[f'H{count}'] = 측정리스트[5].get("time")
                count +=1

        wb.save("C:/Users/유닉솔루션(주)/Desktop/test.xlsx")


        
if __name__ == '__main__' :

    ToExcel()