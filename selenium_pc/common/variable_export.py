# from asyncio.windows_events import NULL
# from multiprocessing.connection import wait
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import variable as var


class Variable():

    def export_excel_variables(self):
        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
        # load_wb = load_workbook("C:/Users/유닉솔루션(주)/Desktop/자동화 엑셀화 데이터 저장 양식_aos.xlsx", data_only=True)
        # 시트 이름으로 불러오기
        load_wb = Workbook()
        load_ws = load_wb.active
        count = 2

        page_name=['common_el','login_el','mainpage_el','search_el','mypage','mobile_el','iptv_el','benefit_el','support_el','direct_el','ujam_el','udoc_el']

        load_ws['B1'] = 'Page Key(페이지명)'
        load_ws['C1'] = 'Element Key(요소명)'
        load_ws['D1'] = 'Variable(요소값)'

        for page in page_name:
            if page in var.__dict__:
                for key,value in var.__dict__.get(page).items():

                    load_ws[f'A{count}'] = count - 1
                    load_ws[f'B{count}'] = page
                    load_ws[f'C{count}'] = key
                    load_ws[f'D{count}'] = value

                    count += 1

        # load_wb.save("C:/Users/유닉솔루션(주)/Desktop/variables.xlsx")
        load_wb.save(f"{os.getcwd()}/variables.xlsx")



    def set_variables(self):
        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
        load_wb = load_workbook(f"{os.getcwd()}\\variables.xlsx", data_only=True)
        # 시트 이름으로 불러오기
        load_ws = load_wb['시트1']


        all_values = {}
        for row in load_ws.rows:
            # A1 셀은 공백으로 패스
            if row[0] == load_ws['A1']:
                print(f"{row[0]} == {load_ws['A1']}")
                continue

            # 데이터 유무(count number)
            if row[0].value:

                page_name =  row[1].value # B1
                key = row[2].value # C1
                value = row[3].value # D1
                assert page_name or key or value, print("일부 데이터 None")

                # {} 초기화
                try:
                    all_values[page_name]
                except:
                    all_values[page_name] = {}

                all_values[page_name][key] = value
            elif row[0].value == row[1].value == row[2].value == row[3].value == None:
                return all_values
            else:
                print(f"{row[0].value}  {row[1].value}  {row[2].value} {row[3].value}")

                assert row[0] or row[1] or row[2] or row[3], print("일부 데이터 비어있음")
            load_wb.close()




if __name__ == '__main__' :

    Variable().export_excel_variables()
