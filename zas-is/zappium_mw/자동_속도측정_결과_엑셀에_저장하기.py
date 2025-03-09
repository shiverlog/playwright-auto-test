import json
import os
import getpass
from openpyxl import load_workbook

cwd = os.getcwd()
if 'test' not in cwd:
    cwd += '\\test\\output'
if 'output' not in cwd:
    cwd += '\\output'
# 수동 속도 측정 결과가 저장된 엑셀 파일
자동_속도측정_파일_경로 =f'{cwd}\\속도측정_결과.json'
# 저장할 엑셀 파일
속도측정_엑셀_경로 =f'C:/Users/{getpass.getuser()}/Desktop/속도측정_결과.xlsx'
#  수동,자동 속도측정 결과를 작성랑 로그 파일
수동자동_결과_로그_경로 =f'{cwd}\\결과.json'


class SetExcel():
    def __init__(self):
        self.export_excel_data()

    def export_excel_data(self):
        속도측정_결과 = {}
        with open(자동_속도측정_파일_경로,encoding='utf-8') as f:
            속도측정_결과=json.load(f)
            f.close()

        속도측정_평균값_결과=[]
        # 속도측정_평균값_결과 저장 형식 : ['MW_LG_1',1.1,1.2,1.3,1.4,1.5,'MW_LG_2',1.1,1.2,1.3,1.4,1.5, ...]

        for 통신사명,통신사값 in 속도측정_결과.items():
            for 페이지명,값 in 통신사값.items():
                # 속도측정_평균값_결과.append(값["page_id"])
                for 측정값 in 값['측정']:
                    속도측정_평균값_결과.append(측정값)
        print(속도측정_평균값_결과)


        # TODO 메인/검색/GNB/모바일요금제/모바일/IPTV/혜택멤버십/고객지원/다이렉트샵/해외로밍
        측정_페이지_수 = 10 
        
        # TODO 페이지별 5회 측정
        측정_횟수=5
        
        wb = load_workbook(속도측정_엑셀_경로,read_only=False, ## 읽기 전용(읽기 전용에 최적화되어 파일을 불러온다)
        data_only=False ## False면 셀안 공식을 가져오고 True면 공식 적용된 값만을 불러온다.
        )
        ws = wb.active
        측정_시작_행= 4
        측정_시작_열='G'
        시간_시작_열='B'

        count=0
        for 통신사 in range(0,3):
            for 페이지 in range(0,측정_페이지_수):
                행=(측정_시작_행+통신사)+(3*페이지)
                print(f"행 = {행}")
                for 회차 in range(0,측정_횟수):
                    print(f"chr(ord(측정_시작_열)+회차) = {chr(ord(측정_시작_열)+회차)}")
                    ws[chr(ord(측정_시작_열)+회차)+str(행)] = 속도측정_평균값_결과[count]['value']
                    ws[chr(ord(시간_시작_열))+str(행)] = 속도측정_평균값_결과[count]['date']
                    count+=1
                

        wb.save(속도측정_엑셀_경로)
    



        
if __name__ == '__main__' :

    SetExcel()