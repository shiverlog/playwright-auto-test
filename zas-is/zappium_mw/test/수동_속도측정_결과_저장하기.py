import json
import getpass, os
from openpyxl import load_workbook
from datetime import datetime, timedelta

import requests
cwd = os.getcwd()
if 'test' not in cwd:
    cwd += '\\test'
# 수동 속도 측정 결과가 저장된 엑셀 파일
수동_속도측정_파일_경로 =f'C:/Users/{getpass.getuser()}/Desktop/속도측정_결과.xlsx'
# 수동,자동 속도측정 결과를 작성랑 로그 파일
수동자동_결과_로그_경로 =f'{cwd}\\output\\결과.json'




class GetExcel():
    def __init__(self):
        res = self.read_excel_data()
        self.로그_저장(res)
        # self.전송()

    def read_excel_data(self):
        속도측정_평균값_결과=[]

        앱_속도측정_페이지_id =[
            ['APP_LG_1',
            'APP_LG_2',
            'APP_LG_3',
            'APP_LG_4',
            'APP_LG_5',
            'APP_LG_6',
            'APP_LG_7',
            'APP_LG_8',
            'APP_LG_9',
            'APP_LG_10'],
            ['APP_SKT_1',
            'APP_SKT_2',
            'APP_SKT_3',
            'APP_SKT_4',
            'APP_SKT_5',
            'APP_SKT_6',
            'APP_SKT_7',
            'APP_SKT_8',
            'APP_SKT_9',
            'APP_SKT_10'],
            ['APP_KT_1',
            'APP_KT_2',
            'APP_KT_3',
            'APP_KT_4',
            'APP_KT_5',
            'APP_KT_6',
            'APP_KT_7',
            'APP_KT_8',
            'APP_KT_9',
            'APP_KT_10']
        ]

        # TODO 메인/검색/GNB/모바일요금제/모바일/IPTV/혜택멤버십/고객지원/다이렉트샵/해외로밍
        측정_페이지_수 = 10 
        
        # TODO 페이지별 10회 측정
        측정_횟수=5

        wb = load_workbook(수동_속도측정_파일_경로,read_only=False, ## 읽기 전용(읽기 전용에 최적화되어 파일을 불러온다)
        data_only=False ## False면 셀안 공식을 가져오고 True면 공식 적용된 값만을 불러온다.
        )
        ws = wb.active
        측정_시작_행= 35
        측정_시작_열='G'
        테스트시간_열='B'
        
        for 통신사 in range(0,3):
            측정_행=측정_시작_행 + 통신사
            for 페이지_id in range(0,측정_페이지_수):
                # 속도측정_평균값_결과.append(앱_속도측정_페이지_id[통신사][페이지_id])
                평균값 = 0.0
                for 회차 in range(0,측정_횟수):
                    평균값 +=float(ws[chr(ord(측정_시작_열)+회차)+str(측정_행)].value)
                print(f"\n {평균값}/{측정_횟수}")
                평균값 = round(평균값/측정_횟수,2)
                속도측정_평균값_결과.append({"code":앱_속도측정_페이지_id[통신사][페이지_id],"value":평균값,"tested_at":ws[테스트시간_열+str(측정_행)].value})
                측정_행+=3

        wb.close()
        print(속도측정_평균값_결과)
        return 속도측정_평균값_결과
    
    def 로그_저장(self,수동_속도측정):
        # 속도측정 결과 저장
        with open(수동자동_결과_로그_경로,'w',encoding='utf-8') as f:
            json.dump(수동_속도측정,f,ensure_ascii=False,indent=4,default=str)
            f.close()

        with open(수동자동_결과_로그_경로,encoding='utf-8') as f:
            _data=json.load(f)
            print("로그저장 후")
            print(_data)
    
    def 날짜(self):
        now = datetime.now()
        # now = datetime.now() - timedelta(days=1)
        return now.strftime('%Y%m%d')
    
    def 전송(self,url:str="https://dcms.uhdcsre.com/dcms/qa"):
        date = self.날짜()
        print(date)
        _data={}
        
        with open(수동자동_결과_로그_경로,encoding='utf-8') as f:
            _data=json.load(f)
            try:
                # 정상 데이터 검증 ------------------------------
                now = datetime.now()
                if True in [True for i in _data if "MW" in str(list(i.values())) and "MW" in str(list(i.values()))]: 
                    raise Exception("MW 측정값이 존재합니다.")
                elif True in [True for i in _data if "APP" in str(list(i.values()))] and True in [True for i in _data if now.strftime('%Y-%m-%d') in str(list(i.values()))]:
                    pass
                else:
                    raise Exception("데이터를 다시 확인해주세요")
                # 정상 데이터 검증 ------------------------------

                # 자동: A / 수동: M
                res = requests.post(url,json=_data,params={"date":date,"div":"M"}) # 500 error
                print(res)
                if res.status_code == 200:
                    print("전송 완료")
                else:
                    print("전송 실패")
                    print(res)
            except Exception as e:
                print(e)
            
            f.close()

        
if __name__ == '__main__' :

    GetExcel()
    # data = GetExcel.read_excel_data()
    # GetExcel.로그_저장(data)
    # GetExcel.전송()
