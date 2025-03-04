from openpyxl import load_workbook
from openpyxl import Workbook
import common.variable as var

if __name__ == '__main__' :
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
    # load_wb = load_workbook("C:/Users/유닉솔루션(주)/Desktop/자동화 엑셀화 데이터 저장 양식_aos.xlsx", data_only=True)
    # 시트 이름으로 불러오기 
    load_wb = Workbook()
    load_ws = load_wb.active
    count =2



    for key,value in var.common_el.items():
        page_name='common_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.login_el.items():
        page_name='login_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.mainpage_el.items():
        page_name='mainpage_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.myinfo_el.items():
        page_name='myinfo_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.chatbot_el.items():
        page_name='chatbot_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.phone_plan_1dep.items():
        page_name='phone_plan_1dep'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.phone_plan_2dep.items():
        page_name='phone_plan_2dep'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.iptv_el.items():
        page_name='iptv_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.mypage.items():
        page_name='mypage'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.mobile_el.items():
        page_name='mobile_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.support_el.items():
        page_name='support_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    # for key,value in var.usim_el.items():
    #     page_name='usim_el'
        
    #     load_ws[f'B{count}'] = page_name
    #     load_ws[f'C{count}'] = key
    #     load_ws[f'D{count}'] = value
    #     count +=1
    for key,value in var.benefit_el.items():
        page_name='benefit_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.direct_el.items():
        page_name='direct_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.uth_el.items():
        page_name='uth_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
    for key,value in var.udoc_el.items():
        page_name='udoc_el'
        
        load_ws[f'B{count}'] = page_name
        load_ws[f'C{count}'] = key
        load_ws[f'D{count}'] = value
        count +=1
   
    load_wb.save("C:/Users/유닉솔루션(주)/Desktop/자동화 엑셀화 데이터_pc.xlsx")
    
